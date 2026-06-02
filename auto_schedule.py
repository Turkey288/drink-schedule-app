"""
飲料店自動排班系統 v23.0 週期判斷修正版
- 每日上班人數：六日一最少 6 人，其餘最少 7 人、最多 8 人
- 每日技術人員：最少 3 人
- 正職/組長/副店長：每兩週 2例 + 2休
- PT：每週 1例 + 1休
- 5/1-5/3 指定休假規則
- 綠圈「○」= 可加班休假：算休假額度，也可計入人力與技術支援
- 空班「空」= 自畫長假超過週期休假額度的緊急空班，算休假日但不計入人力
- 目標：優先讓每位人員在規定週期內修滿假，同時用 ○ 補足人力缺口
"""

import os
import sys

# Windows/Streamlit 編碼修正：避免 PowerShell cp950 無法輸出特殊符號
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass
import glob
import random
import calendar
import shutil
import tempfile
from copy import copy
from datetime import datetime, timedelta
from collections import defaultdict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

random.seed(42)

# ============================
# 基本設定
# ============================
最低上班人數 = 6
平日最低上班人數 = 7
低人力最低上班人數 = 6
最高上班人數 = 8
最低技術人數 = 3
個人連續上限 = {
    "王森弘": 4,
}

年, 月 = 2026, 6
預設輸入檔 = "飲料店排班自動化_身分下拉更新版.xlsx"
輸出檔 = f"自動排班結果_{年}年{月}月_v39_蔣宜蓁PT版.xlsx"

休假符號 = ["例", "休", "○", "空"]
完全休假符號 = ["例", "休"]
可加班休假符號 = "○"
空班符號 = "空"
固定六日休人員 = {"張聰俊"}
特殊PT設定 = {
    "蔣宜蓁": {
        "身份": "PT",
        "生效日": datetime(2026, 6, 21),
        "可上班星期": {4, 5, 6},  # 週五、週六、週日；週五以可晚班表示
        "可早班": False,
        "可晚班": True,
        "搖飲": False,
        "新人": False,
    }
}
跨月週期扣抵 = defaultdict(dict)
匯入提醒列表 = []
前月休假記錄 = defaultdict(dict)
長假日期記錄 = defaultdict(set)
特殊空班日期記錄 = defaultdict(set)

# 5/1-5/3 指定休假規則
特殊三日需休兩天 = ["王森弘", "萬郁茹", "張聰俊", "張佾暄"]
特殊三日需休一天 = ["彭笙祐", "莊絨媗", "陳瑞銘", "鄭竹家"]

def 月份週區塊(年, 月):
    最後日 = calendar.monthrange(年, 月)[1]
    週名 = ["第一週", "第二週", "第三週", "第四週", "第五週", "第六週"]
    色碼 = ["00B050", "00B050", "F8CBAD", "F8CBAD", "FFD966", "FFD966"]
    區塊 = []
    當日 = datetime(年, 月, 1)
    最後 = datetime(年, 月, 最後日)
    idx = 0
    while 當日 <= 最後:
        開始 = 當日
        到週日天數 = 6 - 開始.weekday()
        結束 = min(開始 + timedelta(days=到週日天數), 最後)
        區塊.append((週名[idx], 開始, 結束, 色碼[idx]))
        當日 = 結束 + timedelta(days=1)
        idx += 1
    return 區塊


def 上個月(年, 月):
    if 月 == 1:
        return 年 - 1, 12
    return 年, 月 - 1


def 下個月(年, 月):
    if 月 == 12:
        return 年 + 1, 1
    return 年, 月 + 1


def 建立兩週週期(年, 月):
    週區塊 = 月份週區塊(年, 月)
    週期 = []
    for i in range(0, len(週區塊) - 1, 2):
        開始 = 週區塊[i][1]
        結束 = 週區塊[i + 1][2]
        週期.append({
            "名稱": f"{開始.month}/{開始.day}-{結束.month}/{結束.day}兩週週期（{週區塊[i][0]}+{週區塊[i + 1][0]}）",
            "開始": 開始,
            "結束": 結束,
            "正職例假": 2,
            "正職休假": 2,
            "特殊區間": False,
            "週期說明": f"{週區塊[i][0]} + {週區塊[i + 1][0]}",
        })
    if len(週區塊) % 2 == 1:
        本名稱, 開始, _, _ = 週區塊[-1]
        結束 = 開始 + timedelta(days=13)
        週期.append({
            "名稱": f"{開始.month}/{開始.day}-{結束.month}/{結束.day}下一個兩週週期（{本名稱}+下月延續週）",
            "開始": 開始,
            "結束": 結束,
            "正職例假": 2,
            "正職休假": 2,
            "特殊區間": False,
            "跨月到下月": True,
            "週期說明": f"{本名稱} + 下月延續週",
        })
    return 週期


def 建立PT週期(年, 月):
    週期 = []
    for 名稱, 開始, 結束, _ in 月份週區塊(年, 月):
        跨月到下月 = False
        if (結束 - 開始).days + 1 < 7:
            結束 = 開始 + timedelta(days=6)
            跨月到下月 = True
        週期.append({
            "名稱": f"{開始.month}/{開始.day}-{結束.month}/{結束.day}{名稱}PT週期",
            "開始": 開始,
            "結束": 結束,
            "PT例假": 1,
            "PT休假": 1,
            "特殊區間": False,
            "跨月到下月": 跨月到下月,
        })
    return 週期


def 週期在日期範圍完整嗎(週期, 所有日期):
    if not 所有日期:
        return False
    return 週期["開始"] >= min(所有日期) and 週期["結束"] <= max(所有日期)


def 前月完整兩週週期列表(年, 月):
    週區塊 = 月份週區塊(年, 月)
    start_index = 0
    if 週區塊 and (週區塊[0][2] - 週區塊[0][1]).days + 1 < 7:
        start_index = 1
    週期 = []
    for i in range(start_index, len(週區塊) - 1, 2):
        週期.append((週區塊[i][1], 週區塊[i + 1][2], f"{週區塊[i][0]}+{週區塊[i + 1][0]}"))
    return 週期


週期列表 = 建立兩週週期(年, 月)
週表頭區塊 = 月份週區塊(年, 月)
PT週期列表 = 建立PT週期(年, 月)


def 設定年月(新年, 新月):
    global 年, 月, 輸出檔, 週期列表, 週表頭區塊, PT週期列表
    年, 月 = 新年, 新月
    輸出檔 = f"自動排班結果_{年}年{月}月_v39_蔣宜蓁PT版.xlsx"
    週期列表 = 建立兩週週期(年, 月)
    週表頭區塊 = 月份週區塊(年, 月)
    PT週期列表 = 建立PT週期(年, 月)


# ============================
# 資料讀取
# ============================
def 選擇輸入檔():
    候選 = []
    for pattern in ["115年休假大表*.xlsx", "*115*休假大表*.xlsx", "*休假大表*.xlsx"]:
        候選.extend(glob.glob(pattern))
    候選 = [
        p for p in 候選
        if not os.path.basename(p).startswith("~$")
        and not os.path.basename(p).startswith("自動排班結果")
    ]
    if 候選:
        候選.sort(key=lambda p: os.path.getmtime(p), reverse=True)
        return 候選[0]
    return 預設輸入檔


def 偵測年月(檔案路徑):
    wb = 載入活頁簿(檔案路徑, data_only=False, read_only=True)
    worksheets = [wb.active] + [ws for ws in wb.worksheets if ws.title != wb.active.title]
    for ws in worksheets:
        title = str(ws.title)
        if len(title) == 6 and title.isdigit():
            y = int(title[:4])
            m = int(title[4:])
            if 1 <= m <= 12:
                return y, m
    for ws in worksheets:
        for row in range(1, min(ws.max_row, 15) + 1):
            for col in range(1, ws.max_column + 1):
                v = ws.cell(row=row, column=col).value
                if isinstance(v, datetime):
                    return v.year, v.month
    return 年, 月


def 活頁簿有工作表(檔案路徑, 工作表名稱):
    try:
        wb = 載入活頁簿(檔案路徑, read_only=True)
        return 工作表名稱 in wb.sheetnames
    except Exception:
        return False


def 載入活頁簿(檔案路徑, data_only=False, read_only=False):
    try:
        from openpyxl import load_workbook
        return load_workbook(檔案路徑, data_only=data_only, read_only=read_only)
    except PermissionError:
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        tmp.close()
        shutil.copyfile(檔案路徑, tmp.name)
        from openpyxl import load_workbook
        return load_workbook(tmp.name, data_only=data_only, read_only=read_only)


def 是新人值(value):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return False
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    return text in {"1", "true", "yes", "y", "是", "新人", "v", "✓", "勾"}


def 判斷新人(row_value, 姓名):
    return 是新人值(row_value) or 姓名 == "張佾暄"


def 套用特殊員工設定(員工列表):
    現有 = {e["姓名"]: e for e in 員工列表}
    下一編號 = max([e.get("編號", 0) for e in 員工列表] or [0]) + 1
    for 姓名, 設定 in 特殊PT設定.items():
        if 姓名 not in 現有:
            員工 = {
                "編號": 下一編號,
                "姓名": 姓名,
                "身份": 設定.get("身份", "PT"),
                "搖飲": 設定.get("搖飲", False),
                "可早班": 設定.get("可早班", True),
                "可晚班": 設定.get("可晚班", True),
                "不可同休群組": "一般",
                "新人": 設定.get("新人", False),
            }
            員工列表.append(員工)
            現有[姓名] = 員工
            下一編號 += 1
        else:
            員工 = 現有[姓名]
            員工["身份"] = 設定.get("身份", 員工.get("身份", "PT"))
            員工["搖飲"] = 設定.get("搖飲", 員工.get("搖飲", False))
            員工["可早班"] = 設定.get("可早班", 員工.get("可早班", True))
            員工["可晚班"] = 設定.get("可晚班", 員工.get("可晚班", True))
            員工["新人"] = 設定.get("新人", 員工.get("新人", False))
    return 員工列表


def 排序員工列表(員工列表):
    # 財務人事列保留在表格最下方，避免和門市排班人員混在一起。
    return sorted(員工列表, key=lambda e: (1 if e["姓名"] == "施明君" else 0, e.get("編號", 0)))


def 讀取員工資料(檔案路徑):
    print("[INFO] 讀取員工資料...")
    wb = 載入活頁簿(檔案路徑, data_only=True, read_only=True)
    ws = wb["員工資料庫"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None for v in row):
            continue
        rows.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})

    員工列表 = []
    for row in rows:
        姓名 = str(row.get("姓名", "")).strip()
        if not 姓名 or 姓名 in ["nan", "None"]:
            continue

        員工 = {
            "編號": int(row["員工編號"]) if pd.notna(row.get("員工編號")) else 0,
            "姓名": 姓名,
            "身份": str(row.get("身份", "")).strip() if pd.notna(row.get("身份")) else "",
            "搖飲": bool(row.get("搖飲")) if pd.notna(row.get("搖飲")) else False,
            "可早班": bool(row.get("可早班")) if pd.notna(row.get("可早班")) else False,
            "可晚班": bool(row.get("可晚班")) if pd.notna(row.get("可晚班")) else False,
            "不可同休群組": str(row.get("不可同休群組", "一般")).strip() if pd.notna(row.get("不可同休群組")) else "一般",
            "新人": 判斷新人(row.get("新人"), 姓名),
        }
        員工列表.append(員工)

    print(f"[OK] 讀取到 {len(員工列表)} 位員工")
    return 員工列表


def 讀取大表員工資料(檔案路徑, 年, 月):
    print("[INFO] 從休假大表讀取員工名單...")
    技術名單 = {"王森弘", "古峻燐", "萬郁茹", "彭笙祐", "劉裕平", "莊絨媗"}
    wb = 載入活頁簿(檔案路徑, data_only=False, read_only=True)
    sheet_name = f"{年}{月:02d}"
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.worksheets[0]

    員工列表 = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        編號 = row[0] if len(row) > 0 else None
        姓名 = str(row[1] or "").strip() if len(row) > 1 else ""
        身份 = str(row[2] or "").replace("\n", "").strip() if len(row) > 2 else ""
        if not 姓名 or 姓名 in ["姓名", "上班人數"]:
            continue
        try:
            編號 = int(編號)
        except (TypeError, ValueError):
            continue
        if "PT" in 身份:
            標準身份 = "PT"
        elif 身份 in ["副店長", "組長", "正職"]:
            標準身份 = 身份
        else:
            標準身份 = 身份 or "其他"
        員工列表.append({
            "編號": 編號,
            "姓名": 姓名,
            "身份": 標準身份,
            "搖飲": 姓名 in 技術名單,
            "可早班": True,
            "可晚班": True,
            "不可同休群組": "管理層" if 標準身份 in ["副店長", "組長"] else "一般",
            "新人": 判斷新人(None, 姓名),
        })

    print(f"[OK] 從休假大表讀取到 {len(員工列表)} 位員工")
    return 員工列表


def 解析日期表頭(value, 年, 月):
    if isinstance(value, datetime):
        if value.month == 月:
            return value.replace(year=年)
        return None
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    for fmt in ["%Y/%m/%d", "%Y-%m-%d", "%m/%d", "%m-%d"]:
        try:
            d = datetime.strptime(text, fmt)
            if fmt in ["%m/%d", "%m-%d"]:
                d = d.replace(year=年)
            if d.year == 年 and d.month == 月:
                return d
        except ValueError:
            pass
    return None


def 標準化假別(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if text == "休":
        return "例"
    if text in ["\\", "/", "＼", "╲"]:
        return "休"
    if text in ["例", "○", "空"]:
        return text
    return None


def 是長假符號(value):
    if value is None:
        return False
    return str(value).strip() in ["\\", "/", "＼", "╲"]


def 尋找橫向休假大表(wb, 員工姓名集合, 年, 月):
    最佳 = None
    for ws in wb.worksheets:
        日期欄 = {}
        日期列 = None
        for row in range(1, min(ws.max_row, 15) + 1):
            row_dates = {}
            for col in range(1, ws.max_column + 1):
                d = 解析日期表頭(ws.cell(row=row, column=col).value, 年, 月)
                if d:
                    row_dates[col] = d
            if len(row_dates) >= 7:
                日期欄 = row_dates
                日期列 = row
                break
        if not 日期欄:
            continue

        員工列 = {}
        for row in range((日期列 or 1) + 1, ws.max_row + 1):
            for col in range(1, min(ws.max_column, 8) + 1):
                name = str(ws.cell(row=row, column=col).value or "").strip()
                if name in 員工姓名集合:
                    員工列[name] = row
                    break

        if len(員工列) >= 3:
            score = len(日期欄) + len(員工列)
            if not 最佳 or score > 最佳[0]:
                最佳 = (score, ws.title, 日期欄, 員工列)
    return 最佳


def 套用休假額度轉空班(休假記錄, 固定不可移動, 員工列表, 所有日期):
    for 員工 in 員工列表:
        姓名 = 員工["姓名"]
        if 姓名 not in 休假記錄:
            continue
        已處理日期 = set()
        for 週期 in 員工適用週期列表(員工):
            目標 = 取得週期目標(員工, 週期)
            日期們 = 週期日期列表(所有日期, 週期)
            已處理日期.update(日期們)
            週期申請 = [(d, 休假記錄[姓名][d]) for d in 日期們 if d in 休假記錄[姓名]]
            if not 週期申請:
                continue

            已例 = 已休 = 0
            排序 = sorted(週期申請, key=lambda x: (
                0 if x[0].weekday() in (5, 6) and 姓名 in 固定六日休人員 else 1,
                0 if x[1] == "例" else 1,
                x[0],
            ))
            新值 = {}
            for d, 假別 in 排序:
                if 假別 == "空":
                    新值[d] = "空" if d in 長假日期記錄.get(姓名, set()) or d in 特殊空班日期記錄.get(姓名, set()) else None
                elif 假別 == "例":
                    if 已例 < 目標["例"]:
                        新值[d] = "例"
                        已例 += 1
                    elif 已休 < 目標["休"]:
                        新值[d] = "休"
                        已休 += 1
                    else:
                        新值[d] = "空" if d in 長假日期記錄.get(姓名, set()) or d in 特殊空班日期記錄.get(姓名, set()) else None
                elif 假別 == "休":
                    if 已休 < 目標["休"]:
                        新值[d] = "休"
                        已休 += 1
                    elif 已例 < 目標["例"]:
                        新值[d] = "例"
                        已例 += 1
                    else:
                        新值[d] = "空" if d in 長假日期記錄.get(姓名, set()) or d in 特殊空班日期記錄.get(姓名, set()) else None
                elif 假別 == "○":
                    if 已休 < 目標["休"]:
                        新值[d] = "○"
                        已休 += 1
                    elif 已例 < 目標["例"]:
                        新值[d] = "例"
                        已例 += 1
                    else:
                        新值[d] = "空" if d in 長假日期記錄.get(姓名, set()) or d in 特殊空班日期記錄.get(姓名, set()) else None

            for d, 假別 in 新值.items():
                if 假別 is None:
                    休假記錄[姓名].pop(d, None)
                    固定不可移動[姓名].discard(d)
                    continue
                休假記錄[姓名][d] = 假別
                固定不可移動[姓名].add(d)

        for d in list(休假記錄[姓名].keys()):
            if d not in 已處理日期:
                if d in 長假日期記錄.get(姓名, set()) or d in 特殊空班日期記錄.get(姓名, set()):
                    休假記錄[姓名][d] = "空"
                elif 休假記錄[姓名][d] in ["例", "休", "○", "空"]:
                    休假記錄[姓名].pop(d, None)
                    固定不可移動[姓名].discard(d)
                    continue
                固定不可移動[姓名].add(d)


def 套用固定休假規則(休假記錄, 固定不可移動, 所有日期):
    for 姓名 in 固定六日休人員:
        for d in 所有日期:
            if d.weekday() in (5, 6):
                休假記錄[姓名][d] = 休假記錄[姓名].get(d, "休")
                固定不可移動[姓名].add(d)


def 套用特殊PT空班規則(休假記錄, 固定不可移動, 所有日期):
    for 姓名, 設定 in 特殊PT設定.items():
        生效日 = 設定["生效日"]
        可上班星期 = 設定["可上班星期"]
        for d in 所有日期:
            if d < 生效日 or d.weekday() not in 可上班星期:
                休假記錄[姓名][d] = 空班符號
                固定不可移動[姓名].add(d)
                特殊空班日期記錄[姓名].add(d)


def 讀取指定月份大表原始假別(檔案路徑, 員工列表, 年, 月):
    wb = 載入活頁簿(檔案路徑, data_only=False)
    員工姓名集合 = {e["姓名"] for e in 員工列表}
    result = 尋找橫向休假大表(wb, 員工姓名集合, 年, 月)
    休假記錄 = defaultdict(dict)
    if not result:
        return 休假記錄, None

    _, sheet_name, 日期欄, 員工列 = result
    ws = wb[sheet_name]
    for 姓名, row in 員工列.items():
        for col, d in 日期欄.items():
            假別 = 標準化假別(ws.cell(row=row, column=col).value)
            if 假別:
                休假記錄[姓名][d] = 假別
    return 休假記錄, sheet_name


def 設定跨月週期扣抵(檔案路徑, 員工列表, 年, 月, 訊息列表=None):
    跨月週期扣抵.clear()
    跨月週期們 = [週期 for 週期 in 週期列表 if 週期.get("跨月")]
    if not 跨月週期們:
        return

    前年, 前月 = 上個月(年, 月)
    try:
        前月假別, sheet_name = 讀取指定月份大表原始假別(檔案路徑, 員工列表, 前年, 前月)
    except Exception as e:
        if 訊息列表 is not None:
            訊息列表.append(["提醒", "跨月週期", f"{前年}/{前月}", "休", f"無法讀取前月最後週扣抵：{e}"])
        return

    前月最後週 = 月份週區塊(前年, 前月)[-1]
    _, 前週開始, 前週結束, _ = 前月最後週
    前週日期 = [前週開始 + timedelta(days=i) for i in range((前週結束 - 前週開始).days + 1)]

    for 週期 in 跨月週期們:
        for 員工 in 員工列表:
            if 員工類型(員工) != "正職":
                continue
            統計 = {"例": 0, "休": 0, "○": 0}
            for d in 前週日期:
                假別 = 前月假別.get(員工["姓名"], {}).get(d)
                if 假別 in 統計:
                    統計[假別] += 1
            跨月週期扣抵[週期["名稱"]][員工["姓名"]] = 統計
            if 訊息列表 is not None and sum(統計.values()) < 2:
                訊息列表.append([
                    "提醒",
                    員工["姓名"],
                    週期["名稱"],
                    "休",
                    f"前月最後週({sheet_name or f'{前年}{前月:02d}'})已休 例{統計['例']} 休{統計['休']} ○{統計['○']}，六月第一週會補足跨月週期",
                ])


def 檢查前月最後週期未休滿(檔案路徑, 員工列表, 年, 月, 訊息列表):
    前年, 前月 = 上個月(年, 月)
    try:
        前月假別, sheet_name = 讀取指定月份大表原始假別(檔案路徑, 員工列表, 前年, 前月)
    except Exception as e:
        訊息列表.append(["提醒", "前月檢查", f"{前年}/{前月}", "休", f"無法檢查前月最後週期：{e}"])
        return
    if not 前月假別:
        訊息列表.append(["提醒", "前月檢查", f"{前年}/{前月}", "休", "找不到前月休假大表，未做前月最後週期檢查"])
        return

    完整兩週 = 前月完整兩週週期列表(前年, 前月)
    if 完整兩週:
        開始, 結束, 週說明 = 完整兩週[-1]
        日期們 = [開始 + timedelta(days=i) for i in range((結束 - 開始).days + 1)]
        for 員工 in 員工列表:
            if 員工類型(員工) != "正職":
                continue
            統計 = {"例": 0, "休": 0, "○": 0}
            for d in 日期們:
                假別 = 前月假別.get(員工["姓名"], {}).get(d)
                if 假別 in 統計:
                    統計[假別] += 1
            總休 = sum(統計.values())
            if 總休 < 4 or 統計["例"] < 2:
                訊息列表.append([
                    "提醒",
                    員工["姓名"],
                    f"{前月}/{開始.day}-{前月}/{結束.day}前月最後兩週({週說明})",
                    "休",
                    f"前月最後兩週目標 2例2休，已休 例{統計['例']} 休{統計['休']} ○{統計['○']}；請確認是否補休完畢",
                ])

    前月週區塊 = 月份週區塊(前年, 前月)
    完整週 = [w for w in 前月週區塊 if (w[2] - w[1]).days + 1 == 7]
    if not 完整週:
        return
    週名, 開始, 結束, _ = 完整週[-1]
    日期們 = [開始 + timedelta(days=i) for i in range(7)]
    for 員工 in 員工列表:
        if 員工類型(員工) != "PT":
            continue
        統計 = {"例": 0, "休": 0, "○": 0}
        for d in 日期們:
            假別 = 前月假別.get(員工["姓名"], {}).get(d)
            if 假別 in 統計:
                統計[假別] += 1
        總休 = sum(統計.values())
        if 總休 < 2 or 統計["例"] < 1:
            訊息列表.append([
                "提醒",
                員工["姓名"],
                f"{前月}/{開始.day}-{前月}/{結束.day}前月最後週({週名})",
                "休",
                f"PT前月最後週目標 1例1休，已休 例{統計['例']} 休{統計['休']} ○{統計['○']}；請確認是否補休完畢",
            ])


def 讀取橫向休假大表(檔案路徑, 員工列表, 年, 月):
    wb = 載入活頁簿(檔案路徑, data_only=False)
    員工姓名集合 = {e["姓名"] for e in 員工列表}
    result = 尋找橫向休假大表(wb, 員工姓名集合, 年, 月)
    休假記錄 = defaultdict(dict)
    固定不可移動 = defaultdict(set)
    if not result:
        return 休假記錄, 固定不可移動, None

    _, sheet_name, 日期欄, 員工列 = result
    ws = wb[sheet_name]
    for 姓名, row in 員工列.items():
        for col, d in 日期欄.items():
            raw_value = ws.cell(row=row, column=col).value
            假別 = 標準化假別(raw_value)
            if not 假別:
                continue
            if 是長假符號(raw_value):
                長假日期記錄[姓名].add(d)
            休假記錄[姓名][d] = 假別
            固定不可移動[姓名].add(d)

    所有日期 = 日期範圍(年, 月)
    套用固定休假規則(休假記錄, 固定不可移動, 所有日期)
    套用特殊PT空班規則(休假記錄, 固定不可移動, 所有日期)
    套用休假額度轉空班(休假記錄, 固定不可移動, 員工列表, 所有日期)
    return 休假記錄, 固定不可移動, sheet_name


def 讀取休假申請(檔案路徑, 員工列表, 年, 月):
    print("[INFO] 讀取休假申請...")
    匯入提醒列表.clear()
    前月休假記錄.clear()
    長假日期記錄.clear()
    特殊空班日期記錄.clear()
    前年, 前月 = 上個月(年, 月)
    try:
        原始前月假別, _ = 讀取指定月份大表原始假別(檔案路徑, 員工列表, 前年, 前月)
        for 姓名, 日期假別 in 原始前月假別.items():
            前月休假記錄[姓名].update(日期假別)
    except Exception:
        pass
    設定跨月週期扣抵(檔案路徑, 員工列表, 年, 月, 匯入提醒列表)
    檢查前月最後週期未休滿(檔案路徑, 員工列表, 年, 月, 匯入提醒列表)
    休假記錄, 固定不可移動, sheet_name = 讀取橫向休假大表(檔案路徑, 員工列表, 年, 月)
    if sheet_name:
        print(f"[OK] 已從「{sheet_name}」讀取 {sum(len(v) for v in 休假記錄.values())} 筆六月休假/空班")
        return 休假記錄, 固定不可移動

    try:
        df = pd.read_excel(檔案路徑, sheet_name="休假資料")
        for _, row in df.iterrows():
            姓名 = row.get("員工姓名")
            休假日期 = row.get("休假日期")
            if pd.isna(姓名) or pd.isna(休假日期):
                continue
            if isinstance(休假日期, str):
                for fmt in ["%Y/%m/%d", "%Y-%m-%d", "%m/%d"]:
                    try:
                        if fmt == "%m/%d":
                            休假日期 = datetime.strptime(f"{年}/{休假日期}", "%Y/%m/%d")
                        else:
                            休假日期 = datetime.strptime(休假日期, fmt)
                        break
                    except ValueError:
                        pass
                else:
                    continue
            if 休假日期.year == 年 and 休假日期.month == 月:
                假別 = 標準化假別(row.get("假別")) or "休"
                休假記錄[str(姓名).strip()][休假日期] = 假別
                固定不可移動[str(姓名).strip()].add(休假日期)
        所有日期 = 日期範圍(年, 月)
        套用固定休假規則(休假記錄, 固定不可移動, 所有日期)
        套用特殊PT空班規則(休假記錄, 固定不可移動, 所有日期)
        套用休假額度轉空班(休假記錄, 固定不可移動, 員工列表, 所有日期)
        print(f"[OK] 讀取到 {sum(len(v) for v in 休假記錄.values())} 筆休假申請")
        return 休假記錄, 固定不可移動
    except Exception as e:
        print(f"[WARNING] 無法讀取休假資料：{e}")
        休假記錄 = defaultdict(dict)
        固定不可移動 = defaultdict(set)
        套用固定休假規則(休假記錄, 固定不可移動, 日期範圍(年, 月))
        套用特殊PT空班規則(休假記錄, 固定不可移動, 日期範圍(年, 月))
        return 休假記錄, 固定不可移動


# ============================
# 基礎工具
# ============================
def 日期範圍(年, 月):
    第一天 = datetime(年, 月, 1)
    最後一天 = datetime(年, 月, calendar.monthrange(年, 月)[1])
    return [第一天 + timedelta(days=i) for i in range((最後一天 - 第一天).days + 1)]


def 日期最低上班人數(日期):
    # 週一、週六、週日維持 6 人，其餘平日維持 7 人。
    return 低人力最低上班人數 if 日期.weekday() in (0, 5, 6) else 平日最低上班人數


def 是休假(v):
    return v in 休假符號


def 計入上班(v):
    # 空白 = 正常上班；○ = 原休假但可加班支援，也計入人力；空 = 空班不出勤
    return v == "" or v == 可加班休假符號


def 計算人數(班表, 員工列表, 日期):
    固定上班 = 0
    固定技術 = 0
    綠圈 = 0
    綠圈技術 = 0
    for 員工 in 員工列表:
        if 員工.get("新人"):
            continue
        v = 班表[員工["姓名"]][日期]
        if v == "":
            固定上班 += 1
            if 員工["搖飲"]:
                固定技術 += 1
        elif v == 可加班休假符號:
            綠圈 += 1
            if 員工["搖飲"]:
                綠圈技術 += 1

    技術支援 = min(max(0, 最低技術人數 - 固定技術), 綠圈技術)
    已用綠圈 = 技術支援
    當日最低上班人數 = 日期最低上班人數(日期)
    人數支援 = min(max(0, 當日最低上班人數 - 固定上班 - 技術支援), max(0, 綠圈 - 已用綠圈))
    return 固定上班 + 技術支援 + 人數支援, 固定技術 + 技術支援


def 日期休假數(班表, 員工列表, 日期):
    return sum(1 for e in 員工列表 if 是休假(班表[e["姓名"]][日期]))


def 員工適用週期列表(員工):
    if 員工類型(員工) == "PT":
        特殊週期 = [週期 for 週期 in 週期列表 if 週期.get("特殊區間")]
        return 特殊週期 + PT週期列表
    return 週期列表


def 找週期(日期, 員工=None):
    週期來源 = 員工適用週期列表(員工) if 員工 else 週期列表
    for 週期 in 週期來源:
        if 週期["開始"] <= 日期 <= 週期["結束"]:
            return 週期
    return None


def 員工類型(員工):
    if 員工["身份"] in ["副店長", "組長", "正職"]:
        return "正職"
    if 員工["身份"] == "PT":
        return "PT"
    return "其他"


def 取得週期目標(員工, 週期):
    姓名 = 員工["姓名"]
    if 週期.get("特殊區間"):
        if 姓名 in 特殊三日需休兩天:
            return {"例": 0, "休": 2, "總休": 2}
        if 姓名 in 特殊三日需休一天:
            return {"例": 0, "休": 1, "總休": 1}
        return {"例": 0, "休": 0, "總休": 0}

    類型 = 員工類型(員工)
    if 類型 == "正職":
        if "正職例假" not in 週期:
            return {"例": 0, "休": 0, "總休": 0}
        例目標 = 週期["正職例假"]
        休目標 = 週期["正職休假"]
        if 週期.get("跨月"):
            扣抵 = 跨月週期扣抵.get(週期["名稱"], {}).get(員工["姓名"], {"例": 0, "休": 0, "○": 0})
            例目標 = max(0, 例目標 - 扣抵.get("例", 0))
            已休類 = 扣抵.get("休", 0) + 扣抵.get("○", 0)
            休目標 = max(0, 休目標 - 已休類)
        return {"例": 例目標, "休": 休目標, "總休": 例目標 + 休目標}
    if 類型 == "PT":
        if "PT例假" not in 週期:
            return {"例": 0, "休": 0, "總休": 0}
        特殊設定 = 特殊PT設定.get(姓名)
        if 特殊設定:
            return {"例": 0, "休": 0, "總休": 0}
        return {"例": 週期["PT例假"], "休": 週期["PT休假"], "總休": 週期["PT例假"] + 週期["PT休假"]}
    return {"例": 0, "休": 0, "總休": 0}


def 週期日期列表(所有日期, 週期):
    return [d for d in 所有日期 if 週期["開始"] <= d <= 週期["結束"]]


def 週期已休統計(班表, 姓名, 週期日期):
    例 = sum(1 for d in 週期日期 if 班表[姓名][d] == "例")
    休 = sum(1 for d in 週期日期 if 班表[姓名][d] == "休")
    圈 = sum(1 for d in 週期日期 if 班表[姓名][d] == "○")
    空 = sum(1 for d in 週期日期 if 班表[姓名][d] == "空")
    return {"例": 例, "休": 休, "○": 圈, "空": 空, "總休": 例 + 休 + 圈}


def 會超休嗎(班表, 員工, 日期, 假別, 所有日期):
    週期 = 找週期(日期, 員工)
    if not 週期:
        return True
    目標 = 取得週期目標(員工, 週期)
    日期們 = 週期日期列表(所有日期, 週期)
    統計 = 週期已休統計(班表, 員工["姓名"], 日期們)

    if 目標["總休"] == 0:
        return True
    if 統計["總休"] + 1 > 目標["總休"]:
        return True
    if 假別 == "例" and 統計["例"] + 1 > 目標["例"]:
        return True
    # ○ 算休，但不吃掉「休」上限，只吃總休額度，避免為了支援人力導致修不滿假
    if 假別 == "休" and 統計["休"] + 統計["○"] + 1 > 目標["休"] + max(0, 目標["例"] - 統計["例"]):
        # 一般情況仍允許休，只要總休沒超。這裡不額外擋，避免特殊週期無例假的情境出錯。
        pass
    return False


def 管理層同休嗎(班表, 管理層, 日期, 員工, 假別):
    if 員工.get("不可同休群組") != "管理層":
        return False
    原值 = 班表[員工["姓名"]][日期]
    班表[員工["姓名"]][日期] = 假別
    # 副店長/組長不可同一天休；例、休、○、空都算休假狀態
    休息主管 = [m for m in 管理層 if 是休假(班表[m["姓名"]][日期])]
    班表[員工["姓名"]][日期] = 原值
    return len(休息主管) >= 2


def 可以排假嗎(班表, 員工列表, 管理層, 員工, 日期, 假別, 所有日期, 訊息列表=None, 允許人力不足轉圈=True):
    姓名 = 員工["姓名"]
    if 班表[姓名][日期] != "":
        return False, 假別

    if 會超休嗎(班表, 員工, 日期, 假別, 所有日期):
        return False, 假別

    if 管理層同休嗎(班表, 管理層, 日期, 員工, 假別):
        return False, 假別

    原值 = 班表[姓名][日期]
    班表[姓名][日期] = 假別
    上班, 技術 = 計算人數(班表, 員工列表, 日期)
    班表[姓名][日期] = 原值

    當日最低上班人數 = 日期最低上班人數(日期)
    if 上班 < 當日最低上班人數 or 技術 < 最低技術人數:
        # 若排「休」會讓人力不足，改排「○」：仍算休假，但該員可加班支援
        if 允許人力不足轉圈 and 假別 == "休" and not 會超休嗎(班表, 員工, 日期, "○", 所有日期):
            return True, "○"
        return False, 假別

    return True, 假別


def 分散日期排序(班表, 員工列表, 週期日期):
    排序資料 = []
    for d in 週期日期:
        上班, 技術 = 計算人數(班表, 員工列表, d)
        休假數 = 日期休假數(班表, 員工列表, d)
        # 優先在上班人數多、技術人數多、休假數少的日子排休
        排序資料.append((d, -上班, -技術, 休假數, random.random()))
    排序資料.sort(key=lambda x: (x[1], x[2], x[3], x[4]))
    return [x[0] for x in 排序資料]


def 補假(班表, 員工列表, 管理層, 員工, 週期, 假別, 數量, 所有日期, 訊息列表):
    if 數量 <= 0:
        return 0

    姓名 = 員工["姓名"]
    日期們 = 週期日期列表(所有日期, 週期)
    已排 = 0

    # 先正常補假，若人力會不足會自動改 ○
    for d in 分散日期排序(班表, 員工列表, 日期們):
        if 已排 >= 數量:
            break
        ok, 實際假別 = 可以排假嗎(班表, 員工列表, 管理層, 員工, d, 假別, 所有日期, 訊息列表)
        if ok:
            班表[姓名][d] = 實際假別
            已排 += 1

    if 已排 < 數量:
        訊息列表.append(["未完成", 姓名, 週期["名稱"], 假別, f"需補 {數量} 天，只成功安排 {已排} 天；請人工確認人力或休假規則"])
    return 已排


def 轉休為綠圈補人力(班表, 員工列表, 所有日期, 訊息列表, 固定不可移動=None):
    """若每日人力或技術不足，優先把當天的「休」轉為「○」，讓人員休假但可加班支援。"""
    for d in 所有日期:
        while True:
            上班, 技術 = 計算人數(班表, 員工列表, d)
            當日最低上班人數 = 日期最低上班人數(d)
            if 上班 >= 當日最低上班人數 and 技術 >= 最低技術人數:
                break

            候選 = []
            for e in 員工列表:
                if e.get("新人"):
                    continue
                if 固定不可移動 and d in 固定不可移動.get(e["姓名"], set()):
                    continue
                v = 班表[e["姓名"]][d]
                if v == "休":
                    # 技術不足時優先找技術人員；人數不足則一般人也可
                    分數 = 0
                    if 技術 < 最低技術人數 and e["搖飲"]:
                        分數 -= 100
                    if 上班 < 當日最低上班人數:
                        分數 -= 50
                    分數 += random.random()
                    候選.append((分數, e))

            if not 候選:
                break

            候選.sort(key=lambda x: x[0])
            e = 候選[0][1]
            班表[e["姓名"]][d] = "○"
            訊息列表.append(["調整", e["姓名"], d.strftime("%m/%d"), "○", "為補足人力/技術，將休假改為綠圈可加班休假"])


def 找連續未休區間(班表, 姓名, 所有日期, 上限=5):
    連續 = []
    for d in 所有日期:
        if 是休假(班表[姓名][d]) or 班表[姓名][d] == 空班符號:
            連續 = []
            continue
        連續.append(d)
        if len(連續) > 上限:
            return 連續[-(上限 + 1):]
    return []


def 取得連續上限(姓名):
    return 個人連續上限.get(姓名, 5)


def 是休息或不用上班(v):
    return 是休假(v) or v == 空班符號


def 跨月連續未休區間(班表, 姓名, 所有日期, 上限=5):
    if not 前月休假記錄:
        return []
    前年, 前月 = 上個月(年, 月)
    前月最後週 = 月份週區塊(前年, 前月)[-1]
    開始 = 前月最後週[1]
    結束 = min(所有日期[0] + timedelta(days=6), 所有日期[-1])
    檢查日期 = [開始 + timedelta(days=i) for i in range((結束 - 開始).days + 1)]
    連續 = []
    for d in 檢查日期:
        if d in 班表.get(姓名, {}):
            v = 班表[姓名][d]
        else:
            v = 前月休假記錄.get(姓名, {}).get(d, "")
        if 是休息或不用上班(v):
            連續 = []
            continue
        連續.append(d)
        if len(連續) > 上限 and any(x.month == 月 for x in 連續):
            return 連續[-(上限 + 1):]
    return []


def 修正連續未休(班表, 員工列表, 管理層, 所有日期, 訊息列表, 固定不可移動):
    """把同週期內既有休假移到連續上班區間中，避免為了修滿假又形成連六。"""
    for 員工 in 員工列表:
        姓名 = 員工["姓名"]
        for _ in range(20):
            區間 = 找連續未休區間(班表, 姓名, 所有日期, 取得連續上限(姓名))
            if not 區間:
                break

            已修正 = False
            候選上班日 = [d for d in 區間 if 班表[姓名][d] == ""]
            候選上班日.sort(key=lambda d: (abs(d.day - 區間[len(區間) // 2].day), random.random()))

            for 新休日期 in 候選上班日:
                週期 = 找週期(新休日期, 員工)
                if not 週期:
                    continue

                週期日期 = 週期日期列表(所有日期, 週期)
                可移動舊休 = [
                    d for d in 週期日期
                    if d not in 區間 and d not in 固定不可移動.get(姓名, set()) and 班表[姓名][d] in ["例", "休", "○"]
                ]
                可移動舊休.sort(key=lambda d: (日期休假數(班表, 員工列表, d), random.random()))

                for 舊休日期 in 可移動舊休:
                    舊假別 = 班表[姓名][舊休日期]
                    班表[姓名][舊休日期] = ""
                    舊日上班, _ = 計算人數(班表, 員工列表, 舊休日期)
                    if 舊日上班 > 最高上班人數:
                        班表[姓名][舊休日期] = 舊假別
                        continue
                    ok, 實際假別 = 可以排假嗎(
                        班表,
                        員工列表,
                        管理層,
                        員工,
                        新休日期,
                        "休" if 舊假別 == "○" else 舊假別,
                        所有日期,
                        訊息列表,
                    )
                    if ok:
                        班表[姓名][新休日期] = 實際假別
                        訊息列表.append([
                            "調整",
                            姓名,
                            f"{舊休日期.strftime('%m/%d')}→{新休日期.strftime('%m/%d')}",
                            實際假別,
                            "為平均分布休假，移動同週期休假避免連六",
                        ])
                        已修正 = True
                        break
                    班表[姓名][舊休日期] = 舊假別

                if 已修正:
                    break

            if not 已修正:
                訊息列表.append([
                    "未完成",
                    姓名,
                    f"{區間[0].strftime('%m/%d')}-{區間[-1].strftime('%m/%d')}",
                    "休",
                    "偵測到連六，但同週期沒有可移動休假可在不超休下修正",
                ])
                break


def 降低超過最高人數(班表, 員工列表, 管理層, 所有日期, 訊息列表):
    """若上班人數高於 8，嘗試把空白上班改成休或○，但不能造成超休。"""
    for d in 所有日期:
        上班, 技術 = 計算人數(班表, 員工列表, d)
        if 上班 <= 最高上班人數:
            continue

        候選 = [e for e in 員工列表 if 班表[e["姓名"]][d] == ""]
        random.shuffle(候選)
        for e in 候選:
            if 上班 <= 最高上班人數:
                break
            週期 = 找週期(d, e)
            if not 週期:
                continue
            目標 = 取得週期目標(e, 週期)
            日期們 = 週期日期列表(所有日期, 週期)
            統計 = 週期已休統計(班表, e["姓名"], 日期們)
            if 統計["總休"] >= 目標["總休"]:
                continue
            ok, 實際假別 = 可以排假嗎(班表, 員工列表, 管理層, e, d, "休", 所有日期, 訊息列表, 允許人力不足轉圈=False)
            if ok:
                班表[e["姓名"]][d] = 實際假別
                訊息列表.append(["調整", e["姓名"], d.strftime("%m/%d"), 實際假別, "當日上班人數超過上限，補排休假"])
                上班, 技術 = 計算人數(班表, 員工列表, d)


def 空班分數(員工):
    分數 = 0
    if 員工["身份"] in ["副店長", "組長"]:
        分數 += 80
    if 員工["搖飲"]:
        分數 += 40
    if 員工["身份"] == "PT":
        分數 -= 20
    if 員工["身份"] not in ["副店長", "組長", "正職", "PT"]:
        分數 -= 30
    return 分數 + random.random()


def 可以改空班嗎(班表, 員工列表, 員工, 日期):
    姓名 = 員工["姓名"]
    if 班表[姓名][日期] != "":
        return False
    班表[姓名][日期] = 空班符號
    上班, 技術 = 計算人數(班表, 員工列表, 日期)
    班表[姓名][日期] = ""
    return 上班 >= 日期最低上班人數(日期) and 技術 >= 最低技術人數


def 可以改綠圈嗎(班表, 員工列表, 員工, 日期):
    姓名 = 員工["姓名"]
    if 班表[姓名][日期] != "":
        return False
    週期 = 找週期(日期, 員工)
    if not 週期 or not 週期在日期範圍完整嗎(週期, 日期範圍(年, 月)):
        return False
    目標 = 取得週期目標(員工, 週期)
    日期們 = 週期日期列表(日期範圍(年, 月), 週期)
    統計 = 週期已休統計(班表, 姓名, 日期們)
    return 統計["總休"] < 目標["總休"]


def 用綠圈修正連續(班表, 員工列表, 所有日期, 訊息列表):
    for 員工 in 員工列表:
        姓名 = 員工["姓名"]
        for _ in range(20):
            區間 = 找連續未休區間(班表, 姓名, 所有日期, 取得連續上限(姓名))
            if not 區間:
                break
            候選日期 = [d for d in 區間 if 可以改綠圈嗎(班表, 員工列表, 員工, d)]
            候選日期.sort(key=lambda d: (abs(d.day - 區間[len(區間) // 2].day), random.random()))
            if not 候選日期:
                break
            d = 候選日期[0]
            班表[姓名][d] = "○"
            訊息列表.append(["調整", 姓名, d.strftime("%m/%d"), "○", "以綠圈休假切開連續上班"])


def 用空班修正人數與連續(班表, 員工列表, 所有日期, 訊息列表):
    """空班只處理既有空班與人力安全，不再為一般人員新增空班。"""
    for d in 所有日期:
        while True:
            上班, _ = 計算人數(班表, 員工列表, d)
            if 上班 <= 最高上班人數:
                break
            候選 = [e for e in 員工列表 if 可以改綠圈嗎(班表, 員工列表, e, d)]
            if not 候選:
                break
            候選.sort(key=空班分數)
            e = 候選[0]
            班表[e["姓名"]][d] = "○"
            訊息列表.append(["調整", e["姓名"], d.strftime("%m/%d"), "○", "當日人數超過上限，改排綠圈休假"])


def 用空班修正跨月連續(班表, 員工列表, 所有日期, 固定不可移動, 訊息列表):
    """檢查前月最後週接本月第一週的連六，優先在本月用綠圈切開。"""
    for 員工 in 員工列表:
        姓名 = 員工["姓名"]
        for _ in range(10):
            區間 = 跨月連續未休區間(班表, 姓名, 所有日期, 取得連續上限(姓名))
            if not 區間:
                break
            本月候選 = [
                d for d in 區間
                if d in 班表[姓名] and 可以改綠圈嗎(班表, 員工列表, 員工, d)
            ]
            本月候選.sort(key=lambda d: (abs(d.day - 區間[len(區間) // 2].day), random.random()))
            if not 本月候選:
                訊息列表.append([
                    "未完成",
                    姓名,
                    f"{區間[0].strftime('%m/%d')}-{區間[-1].strftime('%m/%d')}",
                    "○",
                    "跨月偵測到連六，但六月段無法在不超休下排綠圈；空班僅允許長假日期使用",
                ])
                break
            d = 本月候選[0]
            班表[姓名][d] = "○"
            訊息列表.append([
                "調整",
                姓名,
                d.strftime("%m/%d"),
                "○",
                "修正五月最後週接六月第一週連六，排綠圈切開",
            ])


def 移動既有休修正跨月連續(班表, 員工列表, 所有日期, 固定不可移動, 訊息列表):
    """不新增假、不用空班，將本月同週期既有非自畫休往前移來切開跨月連六。"""
    for 員工 in 員工列表:
        姓名 = 員工["姓名"]
        for _ in range(10):
            區間 = 跨月連續未休區間(班表, 姓名, 所有日期, 取得連續上限(姓名))
            if not 區間:
                break
            本月候選 = [d for d in 區間 if d in 班表[姓名] and 班表[姓名][d] == ""]
            本月候選.sort(key=lambda d: (abs(d.day - 區間[len(區間) // 2].day), random.random()))
            已修正 = False
            for 新休日期 in 本月候選:
                週期 = 找週期(新休日期, 員工)
                if not 週期:
                    continue
                週期日期 = 週期日期列表(所有日期, 週期)
                可移動舊休 = [
                    d for d in 週期日期
                    if d not in 區間
                    and d not in 固定不可移動.get(姓名, set())
                    and 班表[姓名][d] in ["例", "休", "○"]
                ]
                可移動舊休.sort(key=lambda d: (日期休假數(班表, 員工列表, d), random.random()))
                for 舊休日期 in 可移動舊休:
                    舊假別 = 班表[姓名][舊休日期]
                    班表[姓名][舊休日期] = ""
                    舊日上班, 舊日技術 = 計算人數(班表, 員工列表, 舊休日期)
                    if 舊日上班 > 最高上班人數 or 舊日技術 < 最低技術人數:
                        班表[姓名][舊休日期] = 舊假別
                        continue
                    新假別 = "○" if 舊假別 == "休" else 舊假別
                    班表[姓名][新休日期] = 新假別
                    新日上班, 新日技術 = 計算人數(班表, 員工列表, 新休日期)
                    if 新日上班 < 日期最低上班人數(新休日期) or 新日技術 < 最低技術人數:
                        班表[姓名][新休日期] = ""
                        班表[姓名][舊休日期] = 舊假別
                        continue
                    訊息列表.append([
                        "調整",
                        姓名,
                        f"{舊休日期.strftime('%m/%d')}→{新休日期.strftime('%m/%d')}",
                        新假別,
                        "移動既有非自畫休，修正五月最後週接六月第一週連六",
                    ])
                    已修正 = True
                    break
                if 已修正:
                    break
            if not 已修正:
                break


def 非自畫休轉綠圈(班表, 員工列表, 所有日期, 固定不可移動, 訊息列表):
    """員工自畫假維持原樣；系統安排的休假統一改為綠圈可加班休假。"""
    for 員工 in 員工列表:
        姓名 = 員工["姓名"]
        固定日期 = 固定不可移動.get(姓名, set())
        for d in 所有日期:
            if d in 固定日期:
                continue
            if 班表[姓名][d] == "休":
                班表[姓名][d] = "○"
                訊息列表.append([
                    "調整",
                    姓名,
                    d.strftime("%m/%d"),
                    "○",
                    "非自畫假之系統補休改為綠圈可加班休假",
                ])


def 檢查每日人力(班表, 員工列表, 所有日期):
    問題 = []
    for d in 所有日期:
        上班, 技術 = 計算人數(班表, 員工列表, d)
        當日最低上班人數 = 日期最低上班人數(d)
        if 上班 < 當日最低上班人數:
            問題.append([d.strftime("%m/%d"), "每日人數不足", f"上班 {上班} 人，低於 {當日最低上班人數} 人"])
        if 上班 > 最高上班人數:
            問題.append([d.strftime("%m/%d"), "每日人數過多", f"上班 {上班} 人，高於 {最高上班人數} 人"])
        if 技術 < 最低技術人數:
            問題.append([d.strftime("%m/%d"), "技術人員不足", f"技術 {技術} 人，低於 {最低技術人數} 人"])
    return 問題


def 檢查週期修假(班表, 員工列表, 所有日期):
    問題 = []
    for e in 員工列表:
        for 週期 in 員工適用週期列表(e):
            日期們 = 週期日期列表(所有日期, 週期)
            目標 = 取得週期目標(e, 週期)
            統計 = 週期已休統計(班表, e["姓名"], 日期們)
            完整週期 = 週期在日期範圍完整嗎(週期, 所有日期)
            if 完整週期 and 統計["總休"] < 目標["總休"]:
                問題.append(["未修滿", e["姓名"], 週期["名稱"], "休", f"目標 {目標['總休']} 天，實際 {統計['總休']} 天"])
            if 統計["總休"] > 目標["總休"]:
                問題.append(["超休", e["姓名"], 週期["名稱"], "休", f"目標 {目標['總休']} 天，實際 {統計['總休']} 天"])
            if not 週期.get("特殊區間") and e["身份"] in ["副店長", "組長", "正職", "PT"]:
                if 完整週期 and 統計["例"] != 目標["例"]:
                    問題.append(["例假不符", e["姓名"], 週期["名稱"], "例", f"目標例假 {目標['例']} 天，實際 {統計['例']} 天"])
    return 問題


def 檢查連續未休(班表, 員工列表, 所有日期):
    問題 = []
    for e in 員工列表:
        上限 = 取得連續上限(e["姓名"])
        區間 = 找連續未休區間(班表, e["姓名"], 所有日期, 上限)
        if 區間:
            問題.append([
                "連續未休",
                e["姓名"],
                f"{區間[0].strftime('%m/%d')}-{區間[-1].strftime('%m/%d')}",
                "休",
                f"超過 {上限} 天未安排休假，請人工確認",
            ])
    return 問題


# ============================
# 排班主流程
# ============================
def 產生班表(員工列表, 休假記錄, 固定不可移動, 年, 月):
    print("[RUN] 開始排班...")
    所有日期 = 日期範圍(年, 月)
    訊息列表 = list(匯入提醒列表)

    班表 = {員工["姓名"]: {d: "" for d in 所有日期} for 員工 in 員工列表}
    管理層 = [e for e in 員工列表 if e["不可同休群組"] == "管理層"]

    # 1. 先處理員工自畫假：休=例、\=休、超額=空班，皆固定不移動
    for 姓名, 日期假別 in 休假記錄.items():
        員工 = next((e for e in 員工列表 if e["姓名"] == 姓名), None)
        if not 員工:
            continue
        for d, 假別 in 日期假別.items():
            if d not in 班表[姓名]:
                continue
            班表[姓名][d] = 假別
            if 假別 == "空":
                if d in 特殊空班日期記錄.get(姓名, set()):
                    訊息列表.append(["空班", 姓名, d.strftime("%m/%d"), "空", "依特殊PT可上班日設定排空班，不計入人力"])
                else:
                    訊息列表.append(["空班", 姓名, d.strftime("%m/%d"), "空", "自畫長假超過週期休假額度，緊急改列空班，算休假日但不計入人力"])

    # 2. 5/1-5/3 指定休假
    特殊週期 = next((週期 for 週期 in 週期列表 if 週期.get("特殊區間")), None)
    if 特殊週期:
        for 姓名 in 特殊三日需休兩天 + 特殊三日需休一天:
            員工 = next((e for e in 員工列表 if e["姓名"] == 姓名), None)
            if not 員工:
                continue
            目標 = 取得週期目標(員工, 特殊週期)
            日期們 = 週期日期列表(所有日期, 特殊週期)
            統計 = 週期已休統計(班表, 姓名, 日期們)
            需補 = max(0, 目標["總休"] - 統計["總休"])
            補假(班表, 員工列表, 管理層, 員工, 特殊週期, "休", 需補, 所有日期, 訊息列表)

    # 3. 正職/組長/副店長兩週週期：2例2休
    for 週期 in [p for p in 週期列表 if not p.get("特殊區間")]:
        if not 週期在日期範圍完整嗎(週期, 所有日期):
            print(f"   記錄 {週期['名稱']}（跨到下月，不在本月硬補滿）")
            continue
        print(f"   處理 {週期['名稱']}")
        日期們 = 週期日期列表(所有日期, 週期)
        for 員工 in 員工列表:
            類型 = 員工類型(員工)
            if 類型 != "正職":
                continue
            目標 = 取得週期目標(員工, 週期)
            統計 = 週期已休統計(班表, 員工["姓名"], 日期們)

            # 先補例假，○ 不可替代例假
            需補例 = max(0, 目標["例"] - 統計["例"])
            補假(班表, 員工列表, 管理層, 員工, 週期, "例", 需補例, 所有日期, 訊息列表)

            統計 = 週期已休統計(班表, 員工["姓名"], 日期們)
            # 再補休假，若人力不足會排成 ○
            需補總休 = max(0, 目標["總休"] - 統計["總休"])
            補假(班表, 員工列表, 管理層, 員工, 週期, "休", 需補總休, 所有日期, 訊息列表)

    # 4. PT每週週期：1例1休
    for 週期 in PT週期列表:
        if not 週期在日期範圍完整嗎(週期, 所有日期):
            print(f"   記錄 {週期['名稱']}（跨到下月，不在本月硬補滿）")
            continue
        print(f"   處理 {週期['名稱']}")
        日期們 = 週期日期列表(所有日期, 週期)
        for 員工 in 員工列表:
            if 員工類型(員工) != "PT":
                continue
            目標 = 取得週期目標(員工, 週期)
            統計 = 週期已休統計(班表, 員工["姓名"], 日期們)

            需補例 = max(0, 目標["例"] - 統計["例"])
            補假(班表, 員工列表, 管理層, 員工, 週期, "例", 需補例, 所有日期, 訊息列表)

            統計 = 週期已休統計(班表, 員工["姓名"], 日期們)
            需補總休 = max(0, 目標["總休"] - 統計["總休"])
            補假(班表, 員工列表, 管理層, 員工, 週期, "休", 需補總休, 所有日期, 訊息列表)

    # 5. 若有人力不足，將休改成 ○ 補人力
    轉休為綠圈補人力(班表, 員工列表, 所有日期, 訊息列表, 固定不可移動)

    # 6. 若人數超過 8，嘗試補休降低人數，但不超休
    降低超過最高人數(班表, 員工列表, 管理層, 所有日期, 訊息列表)

    # 7. 再補一次未修滿的人：依身份週期補足，盡量用 ○ 保住人力
    for 員工 in 員工列表:
        for 週期 in 員工適用週期列表(員工):
            if not 週期在日期範圍完整嗎(週期, 所有日期):
                continue
            日期們 = 週期日期列表(所有日期, 週期)
            目標 = 取得週期目標(員工, 週期)
            統計 = 週期已休統計(班表, 員工["姓名"], 日期們)
            缺 = 目標["總休"] - 統計["總休"]
            if 缺 > 0:
                補假(班表, 員工列表, 管理層, 員工, 週期, "休", 缺, 所有日期, 訊息列表)

    轉休為綠圈補人力(班表, 員工列表, 所有日期, 訊息列表, 固定不可移動)
    修正連續未休(班表, 員工列表, 管理層, 所有日期, 訊息列表, 固定不可移動)
    轉休為綠圈補人力(班表, 員工列表, 所有日期, 訊息列表, 固定不可移動)
    用空班修正人數與連續(班表, 員工列表, 所有日期, 訊息列表)
    轉休為綠圈補人力(班表, 員工列表, 所有日期, 訊息列表, 固定不可移動)
    非自畫休轉綠圈(班表, 員工列表, 所有日期, 固定不可移動, 訊息列表)
    用綠圈修正連續(班表, 員工列表, 所有日期, 訊息列表)
    用空班修正人數與連續(班表, 員工列表, 所有日期, 訊息列表)
    移動既有休修正跨月連續(班表, 員工列表, 所有日期, 固定不可移動, 訊息列表)
    用空班修正跨月連續(班表, 員工列表, 所有日期, 固定不可移動, 訊息列表)
    轉休為綠圈補人力(班表, 員工列表, 所有日期, 訊息列表, 固定不可移動)

    # 最終檢查
    for p in 檢查每日人力(班表, 員工列表, 所有日期):
        訊息列表.append(["違規", "每日人力", p[0], p[1], p[2]])
    訊息列表.extend(檢查週期修假(班表, 員工列表, 所有日期))
    訊息列表.extend(檢查連續未休(班表, 員工列表, 所有日期))

    return 班表, 所有日期, 訊息列表


# ============================
# Excel 輸出
# ============================
def 輸出Excel(班表, 員工列表, 所有日期, 年, 月, 輸出檔, 訊息列表):
    print("[EXPORT] 輸出Excel...")
    wb = Workbook()
    ws = wb.active
    ws.title = f"{年}年{月}月班表"
    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass

    紅底 = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    綠底 = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    淺綠底 = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    灰底 = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    空班藍底 = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
    例假橘底 = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    病假桃底 = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")
    黃底 = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    淺黃底 = PatternFill(start_color="FFF96B", end_color="FFF96B", fill_type="solid")
    藍底 = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    深藍底 = PatternFill(start_color="9DC3E6", end_color="9DC3E6", fill_type="solid")
    粉紫底 = PatternFill(start_color="EE82EE", end_color="EE82EE", fill_type="solid")
    桃底 = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
    淺紅底 = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
    白底 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    細框 = Border(left=Side(style="thin", color="000000"), right=Side(style="thin", color="000000"), top=Side(style="thin", color="000000"), bottom=Side(style="thin", color="000000"))
    中框 = Border(left=Side(style="medium", color="000000"), right=Side(style="medium", color="000000"), top=Side(style="medium", color="000000"), bottom=Side(style="medium", color="000000"))
    綠粗框 = Border(left=Side(style="medium", color="00B050"), right=Side(style="medium", color="00B050"), top=Side(style="medium", color="00B050"), bottom=Side(style="medium", color="00B050"))

    最後欄 = len(所有日期) + 3
    人力輔助欄 = 最後欄 + 1
    技術輔助欄 = 最後欄 + 2
    人力輔助欄字母 = get_column_letter(人力輔助欄)
    技術輔助欄字母 = get_column_letter(技術輔助欄)
    員工起始列 = 5
    員工結束列 = 員工起始列 + len(員工列表) - 1

    # 標題
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=最後欄)
    ws["A1"] = f"得正#竹北博愛計劃{月}月休假大表"
    ws["A1"].font = Font(name="標楷體", size=18, bold=True, color="000000")
    ws["A1"].fill = 白底
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].border = 綠粗框
    ws.row_dimensions[1].height = 28

    # 週期表頭
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    ws.cell(row=2, column=1, value="週期")
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=2, column=1).fill = 藍底

    for 名稱, 開始, 結束, 色碼 in 週表頭區塊:
        start_col = 3 + (開始.day - 1)
        end_col = 3 + (結束.day - 1)
        ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
        c = ws.cell(row=2, column=start_col, value=名稱)
        c.fill = PatternFill(start_color=色碼, end_color=色碼, fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font = Font(name="標楷體", bold=True, color="000000")

    # 日期列
    ws["A3"] = "姓名"
    ws["B3"] = "身份"
    ws["A3"].fill = 藍底
    ws["B3"].fill = 藍底
    for i, d in enumerate(所有日期, start=3):
        cell = ws.cell(row=3, column=i, value=f"{d.month}/{d.day}")
        cell.fill = 淺黃底
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(name="標楷體", size=9, bold=True)

    ws.cell(row=3, column=最後欄, value="休假天數").fill = 黃底

    # 星期列
    ws["B4"] = "星期"
    星期對照 = ["一", "二", "三", "四", "五", "六", "日"]
    for i, d in enumerate(所有日期, start=3):
        cell = ws.cell(row=4, column=i, value=星期對照[d.weekday()])
        cell.alignment = Alignment(horizontal="center")
        cell.fill = 淺黃底
        cell.font = Font(name="標楷體", size=10, bold=True)

    # 員工列
    for r, 員工 in enumerate(員工列表, start=5):
        ws.cell(row=r, column=1, value=員工["姓名"])
        身份顯示 = f'{員工["身份"]}\n新人' if 員工.get("新人") else 員工["身份"]
        ws.cell(row=r, column=2, value=身份顯示)
        ws.cell(row=r, column=人力輔助欄, value=0 if 員工.get("新人") else 1)
        ws.cell(row=r, column=技術輔助欄, value=1 if 員工["搖飲"] and not 員工.get("新人") else 0)
        ws.cell(row=r, column=1).fill = 深藍底
        ws.cell(row=r, column=2).fill = 藍底
        ws.cell(row=r, column=1).font = Font(name="標楷體", bold=True)
        ws.cell(row=r, column=2).font = Font(name="標楷體", bold=True)
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if 員工["身份"] in ["副店長", "組長"]:
            ws.cell(row=r, column=2).fill = 深藍底
        if 員工.get("新人"):
            ws.cell(row=r, column=2).fill = 灰底

        for c, d in enumerate(所有日期, start=3):
            v = 班表[員工["姓名"]][d]
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="標楷體", size=12, bold=True)
            if v == "例":
                cell.fill = 例假橘底
                cell.font = Font(name="標楷體", color="FF0000", bold=True)
            elif v == "休":
                cell.fill = 綠底
                cell.font = Font(name="標楷體", color="000000", bold=True)
            elif v == "○":
                cell.fill = 綠底
                cell.font = Font(name="標楷體", color="000000", bold=True, size=11)
            elif v == "空":
                cell.fill = 空班藍底
                cell.font = Font(name="標楷體", color="000000", bold=True)
            elif d.weekday() >= 5:
                cell.fill = 粉紫底

        first_date_col = get_column_letter(3)
        last_date_col = get_column_letter(最後欄 - 1)
        total_formula = (
            f'=COUNTIF({first_date_col}{r}:{last_date_col}{r},"例")'
            f'+COUNTIF({first_date_col}{r}:{last_date_col}{r},"休")'
            f'+COUNTIF({first_date_col}{r}:{last_date_col}{r},"○")'
            f'+COUNTIF({first_date_col}{r}:{last_date_col}{r},"空")'
        )
        ws.cell(row=r, column=最後欄, value=total_formula).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=最後欄).fill = 白底
        ws.cell(row=r, column=最後欄).font = Font(name="標楷體")

    # 統計列
    總列 = len(員工列表) + 5
    ws.cell(row=總列, column=1, value="上班人數")
    ws.cell(row=總列 + 1, column=1, value="技術人員")
    ws.cell(row=總列 + 2, column=1, value="○ 可加班人數")
    for rr in [總列, 總列 + 1, 總列 + 2]:
        ws.cell(row=rr, column=1).font = Font(name="標楷體", bold=True)
        ws.cell(row=rr, column=1).fill = 黃底

    for i, d in enumerate(所有日期, start=3):
        col_letter = get_column_letter(i)
        day_range = f"{col_letter}${員工起始列}:{col_letter}${員工結束列}"
        people_range = f"${人力輔助欄字母}${員工起始列}:${人力輔助欄字母}${員工結束列}"
        tech_range = f"${技術輔助欄字母}${員工起始列}:${技術輔助欄字母}${員工結束列}"
        fixed_people = f'SUMPRODUCT(--({day_range}=""),{people_range})'
        green_people = f'SUMPRODUCT(--({day_range}="○"),{people_range})'
        fixed_tech = f'SUMPRODUCT(--({day_range}=""),{tech_range})'
        green_tech = f'SUMPRODUCT(--({day_range}="○"),{tech_range})'
        tech_support = f"MIN(MAX(0,{最低技術人數}-{fixed_tech}),{green_tech})"
        當日最低上班人數 = 日期最低上班人數(d)
        people_support = f"MIN(MAX(0,{當日最低上班人數}-{fixed_people}-{tech_support}),MAX(0,{green_people}-{tech_support}))"
        people_formula = f"={fixed_people}+{tech_support}+{people_support}"
        tech_formula = f"={fixed_tech}+{tech_support}"
        overtime_formula = f"={green_people}"
        c1 = ws.cell(row=總列, column=i, value=people_formula)
        c2 = ws.cell(row=總列 + 1, column=i, value=tech_formula)
        c3 = ws.cell(row=總列 + 2, column=i, value=overtime_formula)
        for c in [c1, c2, c3]:
            c.alignment = Alignment(horizontal="center")
            c.fill = 黃底
            c.font = Font(name="標楷體", color="FF0000", bold=True)
        c3.fill = 綠底
        c3.font = Font(color="000000", bold=True)

    # 格式
    for row in ws.iter_rows(min_row=1, max_row=總列 + 2, min_col=1, max_col=最後欄):
        for cell in row:
            cell.border = 細框
            new_font = copy(cell.font)
            new_font.name = "標楷體"
            cell.font = new_font

    # 加粗外框與週別分隔線
    for c in range(1, 最後欄 + 1):
        ws.cell(row=1, column=c).border = 綠粗框
        ws.cell(row=2, column=c).border = 細框
    for r in range(1, 總列 + 3):
        ws.cell(row=r, column=1).border = 細框
        ws.cell(row=r, column=最後欄).border = 細框
    for 名稱, 開始, 結束, 色碼 in 週表頭區塊:
        col = 3 + (開始.day - 1)
        for r in range(2, 總列 + 3):
            old = ws.cell(row=r, column=col).border
            ws.cell(row=r, column=col).border = Border(
                left=Side(style="medium", color="000000"),
                right=old.right,
                top=old.top,
                bottom=old.bottom,
            )

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 9
    for i in range(3, 最後欄 + 1):
        ws.column_dimensions[get_column_letter(i)].width = 4.2
    ws.column_dimensions[人力輔助欄字母].hidden = True
    ws.column_dimensions[技術輔助欄字母].hidden = True
    ws.freeze_panes = "C5"

    # 右側圖例，貼近範例的色塊說明
    legend_col = 最後欄 + 4
    ws.cell(row=7, column=legend_col, value="例").fill = 例假橘底
    ws.cell(row=7, column=legend_col).font = Font(name="標楷體", bold=True, color="FF0000")
    ws.cell(row=7, column=legend_col + 1, value="休").fill = 綠底
    ws.cell(row=7, column=legend_col + 2, value="○").fill = 綠底
    ws.cell(row=10, column=legend_col, value="休").fill = 綠底
    ws.cell(row=10, column=legend_col + 1, value="空").fill = 空班藍底
    ws.cell(row=10, column=legend_col + 2, value="事").fill = 紅底
    ws.cell(row=10, column=legend_col + 3, value="病").fill = 病假桃底
    ws.cell(row=10, column=legend_col + 2).font = Font(name="標楷體", bold=True, color="FFFFFF")
    for row in [7, 10]:
        for col in range(legend_col, legend_col + 4):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if not (row == 7 and col == legend_col) and not (row == 10 and col == legend_col + 2):
                cell.font = Font(name="標楷體", bold=True, color="000000")
            cell.border = 細框
            ws.column_dimensions[get_column_letter(col)].width = 4

    # 主表下方休假規定
    規定列 = 總列 + 4
    ws.merge_cells(start_row=規定列, start_column=1, end_row=規定列 + 4, end_column=最後欄)
    規定文字 = (
        "休假規定：\n"
        "一、每人每月可畫假3日(視為例假)，其餘休假由公司排定，六日一最低上班6人、其他天最低上班7人(不包含外送員)。\n"
        "二、綠色為休假日，綠○有機會至公司加班。\n"
        "三、臨時請假，請與夥伴協商休假日(綠色)調假，雙方同意後，向經理回報。\n"
        "四、未經同意私自調假，視同曠職。"
    )
    rule_cell = ws.cell(row=規定列, column=1, value=規定文字)
    rule_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    rule_cell.font = Font(name="標楷體", size=11, bold=True)
    rule_cell.fill = 白底
    for row in ws.iter_rows(min_row=規定列, max_row=規定列 + 4, min_col=1, max_col=最後欄):
        for cell in row:
            cell.border = 細框
    for rr in range(規定列, 規定列 + 5):
        ws.row_dimensions[rr].height = 22

    # 違規檢查表
    ws2 = wb.create_sheet("違規檢查")
    ws2.append(["項目", "姓名/類型", "日期/週期", "假別", "說明"])
    for cell in ws2[1]:
        cell.fill = 黃底
        cell.font = Font(bold=True)

    if 訊息列表:
        # 去除完全重複訊息
        seen = set()
        for row in 訊息列表:
            key = tuple(row)
            if key in seen:
                continue
            seen.add(key)
            ws2.append(row)
    else:
        ws2.append(["OK", "-", "-", "-", "未發現違規"])

    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 24
    ws2.column_dimensions["D"].width = 10
    ws2.column_dimensions["E"].width = 80

    # 規則說明表
    ws3 = wb.create_sheet("規則說明")
    說明 = [
        ["符號", "意思", "是否算休假", "是否計入人力"],
        ["空白", "正常上班", "否", "是"],
        ["例", "例假，完全休假", "是", "否"],
        ["休", "員工自畫長假或固定休假，完全休假", "是", "否"],
        ["○", "非自畫假之系統補休：可加班休假，可支援補人力", "是", "是"],
        ["空", "空班：僅限自畫長假超過額度時緊急使用", "是", "否"],
    ]
    for row in 說明:
        ws3.append(row)
    for cell in ws3[1]:
        cell.fill = 黃底
        cell.font = Font(bold=True)
    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 46
    ws3.column_dimensions["C"].width = 16
    ws3.column_dimensions["D"].width = 16

    ws4 = wb.create_sheet("週期對照")
    ws4.append(["類型", "週期", "日期範圍", "週別對照", "規則"])
    for cell in ws4[1]:
        cell.fill = 黃底
        cell.font = Font(bold=True)
    for 週期 in 週期列表:
        ws4.append([
            "正職/組長/副店長",
            週期["名稱"],
            f"{週期['開始'].strftime('%m/%d')}-{週期['結束'].strftime('%m/%d')}",
            週期.get("週期說明", ""),
            "2例 + 2休；跨到下月的週期只記錄本月已畫假，不在本月硬補滿",
        ])
    for 週期 in PT週期列表:
        ws4.append([
            "PT",
            週期["名稱"],
            f"{週期['開始'].strftime('%m/%d')}-{週期['結束'].strftime('%m/%d')}",
            週期["名稱"].split("PT週期")[0],
            "每週 1例 + 1休；只有自畫長假超額才轉空班",
        ])
    for col, width in {"A": 18, "B": 38, "C": 18, "D": 34, "E": 50}.items():
        ws4.column_dimensions[col].width = width

    try:
        wb.save(輸出檔)
        print(f"[OK] 已儲存：{輸出檔}")
    except PermissionError:
        base, ext = os.path.splitext(輸出檔)
        備用輸出檔 = f"{base}_新版_{datetime.now().strftime('%H%M%S')}{ext}"
        wb.save(備用輸出檔)
        print(f"[WARNING] 原輸出檔被開啟或鎖定，已另存：{備用輸出檔}")


# ============================
# 主程式
# ============================
def 執行排班(輸入檔=None, 自動偵測年月=True):
    if not 輸入檔:
        輸入檔 = 選擇輸入檔()
    if 自動偵測年月 and os.path.exists(輸入檔):
        偵測年, 偵測月 = 偵測年月(輸入檔)
        設定年月(偵測年, 偵測月)

    print("=" * 60)
    print("飲料店自動排班系統 v39.0 蔣宜蓁PT版")
    print(f"   每日上班人數：六日一至少 {低人力最低上班人數} 人，其他天至少 {平日最低上班人數} 人，最多 {最高上班人數} 人")
    print(f"   每日技術人數：至少 {最低技術人數} 人")
    print("   新人 = 可排在班表中，但不計入上班人數、技術人員與可加班人數")
    print("   ○ = 算休假，也可加班支援人力")
    print("   空 = 長假超額緊急空班，算休假日但不計入人力")
    print("   蔣宜蓁 = 6/21起週五晚班、六日整天可上班，其餘日期空班")
    print("=" * 60)

    print(f"   使用輸入檔：{輸入檔}")
    if not os.path.exists(輸入檔):
        print(f"[ERROR] 找不到檔案：{輸入檔}")
        print("請確認 py 檔和 Excel 檔放在同一個資料夾。")
        return

    員工資料檔 = 輸入檔 if 活頁簿有工作表(輸入檔, "員工資料庫") else 預設輸入檔
    if 員工資料檔 != 輸入檔:
        print(f"   員工資料來源：{員工資料檔}")

    try:
        員工 = 讀取員工資料(員工資料檔)
    except Exception as e:
        print(f"[WARNING] 員工資料庫讀取失敗，改讀休假大表名單：{e}")
        員工 = 讀取大表員工資料(輸入檔, 年, 月)
    員工 = 排序員工列表(套用特殊員工設定(員工))
    休假, 固定不可移動 = 讀取休假申請(輸入檔, 員工, 年, 月)
    班表, 日期列表, 訊息列表 = 產生班表(員工, 休假, 固定不可移動, 年, 月)

    print("\n[SUMMARY] 最終休假統計：")
    for e in 員工:
        例數 = sum(1 for d in 日期列表 if 班表[e["姓名"]][d] == "例")
        休數 = sum(1 for d in 日期列表 if 班表[e["姓名"]][d] == "休")
        圈數 = sum(1 for d in 日期列表 if 班表[e["姓名"]][d] == "○")
        空數 = sum(1 for d in 日期列表 if 班表[e["姓名"]][d] == "空")
        print(f"  {e['姓名']}({e['身份']}): 例{例數} 休{休數} ○{圈數} 空{空數} 總休{例數+休數+圈數}")

    輸出Excel(班表, 員工, 日期列表, 年, 月, 輸出檔, 訊息列表)

    每日問題 = 檢查每日人力(班表, 員工, 日期列表)
    週期問題 = 檢查週期修假(班表, 員工, 日期列表)
    連續問題 = 檢查連續未休(班表, 員工, 日期列表)
    if 每日問題 or 週期問題 or 連續問題:
        print("\n[WARNING] 仍有問題，請查看 Excel 的『違規檢查』工作表。")
    else:
        print("\n[OK] 檢查通過：每日人力符合，每位人員皆依週期修滿假、未超休，且無連六。")
    print("[DONE] 完成！")
    return 輸出檔


def main():
    輸入檔 = sys.argv[1] if len(sys.argv) > 1 else None
    try:
        執行排班(輸入檔)
    except Exception as e:
        print(f"\n執行失敗：{e}")
        raise


if __name__ == "__main__":
    main()
